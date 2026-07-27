# -*- coding: utf-8 -*-
"""
Generate products.js from SEAGM Excel files.
- Cards list + Video on Demand
- 1 USD = 46 MRU
- No duplicate products
"""

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

CARDS_XLSX = Path(r"c:\Users\pc\Downloads\SEAGM_Card_Products_List.xlsx")
VOD_XLSX = Path(r"c:\Users\pc\Downloads\SEAGM_Video_on_Demand_Prices_2026-07-27.xlsx")
OUT_PATH = Path(__file__).resolve().parent.parent / "products.js"
RATE = 46
WHATSAPP = "22248650585"

CATEGORY_META_CARDS = {
    "Gift Cards": {
        "id": "gift-cards",
        "name_ar": "بطاقات الهدايا",
        "icon": "fa-solid fa-gift",
        "order": 2,
    },
    "Game Cards": {
        "id": "game-cards",
        "name_ar": "بطاقات الألعاب",
        "icon": "fa-solid fa-gamepad",
        "order": 3,
    },
    "Mobile Game Cards": {
        "id": "mobile-game-cards",
        "name_ar": "ألعاب الموبايل",
        "icon": "fa-solid fa-mobile-screen",
        "order": 4,
    },
    "Console & Others": {
        "id": "console-others",
        "name_ar": "كونسول وغيرها",
        "icon": "fa-solid fa-tv",
        "order": 5,
    },
}

VOD_CATEGORY = {
    "id": "video-on-demand",
    "name_ar": "بث وفيديو",
    "icon": "fa-solid fa-clapperboard",
    "order": 6,
}

# brand keyword → (domain, icon, color)
BRAND_MAP = [
    (("itunes", "apple gift", "apple"), "apple.com", "fa-brands fa-apple", "#A2AAAD"),
    (("google play",), "play.google.com", "fa-brands fa-google-play", "#34A853"),
    (("steam",), "store.steampowered.com", "fa-brands fa-steam", "#1B2838"),
    (("playstation", "psn"), "playstation.com", "fa-brands fa-playstation", "#003087"),
    (("xbox",), "xbox.com", "fa-brands fa-xbox", "#107C10"),
    (("nintendo", "eshop"), "nintendo.com", "fa-solid fa-gamepad", "#E60012"),
    (("spotify",), "spotify.com", "fa-brands fa-spotify", "#1DB954"),
    (("netflix",), "netflix.com", "fa-solid fa-film", "#E50914"),
    (("hulu",), "hulu.com", "fa-solid fa-tv", "#1CE783"),
    (("paramount",), "paramountplus.com", "fa-solid fa-film", "#0064FF"),
    (("twitch",), "twitch.tv", "fa-brands fa-twitch", "#9146FF"),
    (("bilibili",), "bilibili.com", "fa-solid fa-play", "#00A1D6"),
    (("iqiyi", "iqi yi"), "iqiyi.com", "fa-solid fa-play", "#00CC36"),
    (("shahid", "mbc"), "shahid.mbc.net", "fa-solid fa-tv", "#E31C23"),
    (("osn",), "osn.com", "fa-solid fa-satellite-dish", "#6B2D8B"),
    (("starzplay", "starz"), "starzplay.com", "fa-solid fa-film", "#000000"),
    (("wetv",), "wetv.vip", "fa-solid fa-film", "#FF6600"),
    (("viu",), "viu.com", "fa-solid fa-play", "#FF6A00"),
    (("sling",), "sling.com", "fa-solid fa-satellite", "#FF6B00"),
    (("showtime",), "showtime.com", "fa-solid fa-film", "#FF0000"),
    (("yango",), "play.yango.com", "fa-solid fa-play", "#FC3F1D"),
    (("tvb",), "tvb.com", "fa-solid fa-tv", "#E60012"),
    (("rtl",), "rtlplus.de", "fa-solid fa-tv", "#DE0000"),
    (("bstation",), "bstation.com", "fa-solid fa-play", "#00A1D6"),
    (("astro", "njoi"), "astro.com.my", "fa-solid fa-satellite-dish", "#E60012"),
    (("disney",), "disneyplus.com", "fa-solid fa-film", "#113CCF"),
    (("hbo",), "hbomax.com", "fa-solid fa-film", "#B10DC9"),
    (("roblox", "rbx"), "roblox.com", "fa-solid fa-cube", "#E2231A"),
    (("discord",), "discord.com", "fa-brands fa-discord", "#5865F2"),
    (("riot", "league of legends", "valorant"), "riotgames.com", "fa-solid fa-trophy", "#D13639"),
    (("free fire", "garena"), "ff.garena.com", "fa-solid fa-fire", "#FF6B00"),
    (("mobile legends",), "mobilelegends.com", "fa-solid fa-shield-halved", "#1E3A8A"),
    (("pubg",), "pubg.com", "fa-solid fa-gun", "#F2A900"),
    (("battle.net", "blizzard"), "battle.net", "fa-brands fa-battle-net", "#00AEFF"),
    (("amazon",), "amazon.com", "fa-brands fa-amazon", "#FF9900"),
    (("razer",), "razer.com", "fa-solid fa-mouse", "#44D62C"),
    (("telegram",), "telegram.org", "fa-brands fa-telegram", "#26A5E4"),
    (("whatsapp",), "whatsapp.com", "fa-brands fa-whatsapp", "#25D366"),
    (("tiktok",), "tiktok.com", "fa-brands fa-tiktok", "#000000"),
    (("youtube",), "youtube.com", "fa-brands fa-youtube", "#FF0000"),
    (("meta quest", "meta"), "meta.com", "fa-brands fa-meta", "#0668E1"),
    (("facebook",), "facebook.com", "fa-brands fa-facebook", "#1877F2"),
    (("snapchat",), "snapchat.com", "fa-brands fa-snapchat", "#FFFC00"),
    (("ea sports", "ea fc", "origin"), "ea.com", "fa-solid fa-futbol", "#000000"),
    (("honor of kings",), "honorofkings.com", "fa-solid fa-crown", "#C9A227"),
    (("genshin", "hoyoverse", "mihoyo"), "genshin.hoyoverse.com", "fa-solid fa-star", "#4A90D9"),
    (("unipin",), "unipin.com", "fa-solid fa-wallet", "#FF6A00"),
    (("eneba",), "eneba.com", "fa-solid fa-tag", "#7B2CBF"),
    (("imvu",), "imvu.com", "fa-solid fa-user", "#FF3399"),
    (("netease",), "neteasegames.com", "fa-solid fa-gamepad", "#E60012"),
    (("point blank",), "pointblank.id", "fa-solid fa-crosshairs", "#FF4500"),
    (("ongame",), "ongame.net", "fa-solid fa-coins", "#F5A623"),
    (("tng", "touch 'n go", "touch n go"), "tngdigital.com.my", "fa-solid fa-wallet", "#0066B3"),
    (("crypto", "binance", "gift me crypto"), "binance.com", "fa-brands fa-bitcoin", "#F3BA2F"),
    (("fortnite", "epic games"), "epicgames.com", "fa-solid fa-gamepad", "#2F2F2F"),
    (("minecraft",), "minecraft.net", "fa-solid fa-cube", "#62B47A"),
    (("clash of clans", "supercell", "clash royale", "brawl stars"), "supercell.com", "fa-solid fa-shield-halved", "#F7C948"),
    (("s sport",), "ssportplus.com", "fa-solid fa-futbol", "#E30613"),
    (("indieflix",), "indieflix.com", "fa-solid fa-film", "#6C63FF"),
]


def logo_url(domain: str) -> str:
    return f"https://logo.clearbit.com/{domain}"


def resolve_brand(name: str):
    n = name.lower()
    for keys, domain, icon, color in BRAND_MAP:
        if any(k in n for k in keys):
            return {
                "domain": domain,
                "icon": icon,
                "color": color,
                "image": logo_url(domain),
            }
    return {
        "domain": None,
        "icon": "fa-solid fa-credit-card",
        "color": "#147BFE",
        "image": None,
    }


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:70] or "product"


def normalize_key(name: str) -> str:
    """Canonical key for duplicate detection."""
    n = name.lower().strip()
    n = n.replace("&", " and ")
    n = re.sub(r"[\[\](){}]", " ", n)
    n = re.sub(r"[^a-z0-9\s+]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    # unify region/currency labels
    n = re.sub(r"\busd\b", "us", n)
    n = re.sub(r"\beur\b", "eu", n)
    n = re.sub(r"\bunited states\b", "us", n)
    n = re.sub(r"\bunited kingdom\b", "uk", n)
    n = re.sub(r"\bgift card\b", "giftcard", n)
    n = re.sub(r"\s+", "", n)
    return n


def parse_sales_score(sales):
    if not sales or str(sales).strip() in ("-", "", "None"):
        return 0
    s = str(sales).upper().replace("+", "").replace(",", "").strip()
    try:
        if "M" in s:
            return float(s.replace("M", "")) * 1_000_000
        if "K" in s:
            return float(s.replace("K", "")) * 1_000
        return float(s)
    except ValueError:
        return 0


def parse_sales_badge(sales, rating):
    score = parse_sales_score(sales)
    if score >= 50_000:
        return "الأكثر مبيعًا"
    if rating is not None:
        try:
            if float(rating) >= 4.99 and score >= 1000:
                return "تقييم ممتاز"
        except (TypeError, ValueError):
            pass
    if score >= 5000:
        return "شائع"
    return None


def usd_options(amounts: list[int | float]) -> list[dict]:
    opts = []
    for d in amounts:
        d_num = float(d)
        # keep integer dollars as int labels
        if d_num == int(d_num):
            d_int = int(d_num)
            label = f"${d_int}"
            duration = f"${d_int} USD"
            oid = f"usd{d_int}"
            usd_val = d_int
        else:
            label = f"${d_num:g}"
            duration = f"${d_num:g} USD"
            oid = f"usd{str(d_num).replace('.', '_')}"
            usd_val = d_num
        mru = int(round(d_num * RATE))
        opts.append(
            {
                "id": oid,
                "label": label,
                "duration": duration,
                "price": mru,
                "usd": usd_val,
            }
        )
    return opts


def denominations_for_cards(name: str) -> list[int]:
    n = name.lower()
    if any(
        k in n
        for k in [
            "free fire",
            "mobile legends",
            "honor of kings",
            "garena",
            "diamonds",
            "token redeem",
            "ea sports fc",
        ]
    ):
        return [1, 2, 5, 10, 20, 50]
    if any(k in n for k in ["roblox", "rbx"]):
        return [10, 25, 50, 100]
    if any(
        k in n
        for k in ["steam", "playstation", "xbox", "nintendo", "battle.net", "riot"]
    ):
        return [5, 10, 20, 50, 100]
    if any(k in n for k in ["itunes", "apple", "google play", "spotify", "netflix"]):
        return [5, 10, 15, 25, 50, 100]
    if any(k in n for k in ["unipin", "razer", "pubg", "genshin"]):
        return [5, 10, 20, 50]
    return [5, 10, 25, 50, 100]


def options_for_vod(name: str) -> list[dict]:
    """Options for streaming/VOD products — values in USD face then × RATE."""
    n = name.lower()

    # Twitch gift cards
    if "twitch" in n:
        return usd_options([20, 25, 50, 100, 200])

    # Netflix / Hulu / Paramount / Sling / Showtime / RTL gift cards
    if any(
        k in n
        for k in [
            "netflix",
            "hulu",
            "paramount",
            "sling",
            "showtime",
            "rtl",
            "gift card",
        ]
    ) and "membership" not in n and "subscription" not in n and "premium" not in n and "vip" not in n:
        return usd_options([10, 15, 25, 50, 100])

    # Bilibili / iQiyi / Viu / WeTV / Bstation memberships
    if any(
        k in n
        for k in [
            "bilibili",
            "iqiyi",
            "viu",
            "wetv",
            "bstation",
            "membership",
            "vip",
            "premium",
        ]
    ):
        # approximate subscription tiers in USD face value
        return [
            {
                "id": "1m",
                "label": "شهر",
                "duration": "شهر",
                "price": int(round(3 * RATE)),
                "usd": 3,
            },
            {
                "id": "3m",
                "label": "3 أشهر",
                "duration": "3 أشهر",
                "price": int(round(8 * RATE)),
                "usd": 8,
            },
            {
                "id": "6m",
                "label": "6 أشهر",
                "duration": "6 أشهر",
                "price": int(round(14 * RATE)),
                "usd": 14,
            },
            {
                "id": "1y",
                "label": "سنة",
                "duration": "سنة",
                "price": int(round(25 * RATE)),
                "usd": 25,
            },
        ]

    # Shahid / OSN+ / StarzPlay / Yango / S Sport / TVB / IndieFlix / NJOI
    if any(
        k in n
        for k in [
            "shahid",
            "osn",
            "starz",
            "yango",
            "s sport",
            "tvb",
            "indieflix",
            "njoi",
            "subscription",
        ]
    ):
        return [
            {
                "id": "1m",
                "label": "شهر",
                "duration": "شهر",
                "price": int(round(5 * RATE)),
                "usd": 5,
            },
            {
                "id": "3m",
                "label": "3 أشهر",
                "duration": "3 أشهر",
                "price": int(round(12 * RATE)),
                "usd": 12,
            },
            {
                "id": "6m",
                "label": "6 أشهر",
                "duration": "6 أشهر",
                "price": int(round(22 * RATE)),
                "usd": 22,
            },
            {
                "id": "1y",
                "label": "سنة",
                "duration": "سنة",
                "price": int(round(40 * RATE)),
                "usd": 40,
            },
        ]

    # default gift-like
    return usd_options([10, 15, 25, 50, 100])


def unique_id(base: str, seen_ids: set) -> str:
    pid = base
    n = 2
    while pid in seen_ids:
        pid = f"{base}-{n}"
        n += 1
    seen_ids.add(pid)
    return pid


def make_product(
    *,
    name: str,
    region: str,
    sales,
    rating,
    category: str,
    category_name: str,
    options: list[dict],
    seen_ids: set,
    seen_keys: set,
    force_featured: bool = False,
) -> dict | None:
    key = normalize_key(name)
    if key in seen_keys:
        return None
    seen_keys.add(key)

    brand = resolve_brand(name)
    pop = parse_sales_score(sales)
    try:
        rating_f = float(rating) if rating not in (None, "-", "") else None
    except (TypeError, ValueError):
        rating_f = None
    badge = parse_sales_badge(sales, rating_f)
    pid = unique_id(slugify(name), seen_ids)

    return {
        "id": pid,
        "name": name,
        "description": f"{category_name} · {region}",
        "icon": brand["icon"],
        "image": brand["image"],
        "color": brand["color"],
        "badge": badge,
        "category": category,
        "categoryName": category_name,
        "region": region,
        "featured": force_featured or bool(badge) or pop >= 5_000,
        "popularity": pop,
        "options": options,
    }


def read_sheet_products(ws):
    """Yield (name, region, sales, rating) from a standard SEAGM product sheet."""
    rows = list(ws.iter_rows(values_only=True))
    start = 0
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == "#" and row[1]:
            start = i + 1
            break
    for row in rows[start:]:
        if not row or row[1] is None:
            continue
        cells = list(row) + [None] * 6
        _num, name, region, sales, rating = cells[:5]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        region = str(region).strip() if region else "Global"
        if region in ("-", "Tunisia?"):
            if "TU" in name or "tunis" in name.lower():
                region = "Tunisia"
            elif region == "-":
                region = "Global"
        sales_s = str(sales).strip() if sales is not None else None
        if sales_s in ("-", "None"):
            sales_s = None
        yield name, region, sales_s, rating


def main():
    products = []
    seen_ids = set()
    seen_keys = set()
    skipped_dupes = []

    # --- Snapchat Plus (fixed store pricing) ---
    products.append(
        {
            "id": "snapchat-plus",
            "name": "Snapchat Plus",
            "description": "اشتراك Snapchat Plus الرسمي — مميزات حصرية وتجربة محسّنة",
            "icon": "fa-brands fa-snapchat",
            "image": logo_url("snapchat.com"),
            "color": "#FFFC00",
            "badge": "الأكثر طلبًا",
            "category": "subscriptions",
            "categoryName": "الاشتراكات",
            "region": "Global",
            "featured": True,
            "popularity": 10_000_000,
            "options": [
                {"id": "3m", "label": "3 أشهر", "duration": "3 أشهر", "price": 190, "usd": None},
                {"id": "6m", "label": "6 أشهر", "duration": "6 أشهر", "price": 370, "usd": None},
                {"id": "1y", "label": "سنة", "duration": "سنة", "price": 730, "usd": None},
            ],
        }
    )
    seen_ids.add("snapchat-plus")
    seen_keys.add(normalize_key("Snapchat Plus"))

    # --- Card products ---
    if CARDS_XLSX.exists():
        wb = openpyxl.load_workbook(CARDS_XLSX, data_only=True)
        for sheet_name, meta in CATEGORY_META_CARDS.items():
            if sheet_name not in wb.sheetnames:
                continue
            for name, region, sales, rating in read_sheet_products(wb[sheet_name]):
                p = make_product(
                    name=name,
                    region=region,
                    sales=sales,
                    rating=rating,
                    category=meta["id"],
                    category_name=meta["name_ar"],
                    options=usd_options(denominations_for_cards(name)),
                    seen_ids=seen_ids,
                    seen_keys=seen_keys,
                )
                if p:
                    products.append(p)
                else:
                    skipped_dupes.append(name)
    else:
        print(f"WARNING: missing {CARDS_XLSX}")

    # --- Video on Demand products ---
    vod_added = 0
    if VOD_XLSX.exists():
        vwb = openpyxl.load_workbook(VOD_XLSX, data_only=True)
        sheet = "قائمة المنتجات الكاملة"
        if sheet in vwb.sheetnames:
            for name, region, sales, rating in read_sheet_products(vwb[sheet]):
                p = make_product(
                    name=name,
                    region=region,
                    sales=sales,
                    rating=rating,
                    category=VOD_CATEGORY["id"],
                    category_name=VOD_CATEGORY["name_ar"],
                    options=options_for_vod(name),
                    seen_ids=seen_ids,
                    seen_keys=seen_keys,
                    force_featured=parse_sales_score(sales) >= 200,
                )
                if p:
                    products.append(p)
                    vod_added += 1
                else:
                    skipped_dupes.append(f"[VOD skip] {name}")
        else:
            print(f"WARNING: sheet {sheet} not found in VOD file")
    else:
        print(f"WARNING: missing {VOD_XLSX}")

    # Sort: featured → category order → popularity → name
    cat_order = {
        "subscriptions": 1,
        "video-on-demand": 2,
        "gift-cards": 3,
        "game-cards": 4,
        "mobile-game-cards": 5,
        "console-others": 6,
    }
    products.sort(
        key=lambda p: (
            0 if p.get("featured") else 1,
            cat_order.get(p.get("category"), 99),
            -int(p.get("popularity") or 0),
            p.get("name", "").lower(),
        )
    )

    categories = [
        {"id": "all", "name": "الكل", "icon": "fa-solid fa-border-all"},
        {"id": "subscriptions", "name": "الاشتراكات", "icon": "fa-solid fa-star"},
        {
            "id": VOD_CATEGORY["id"],
            "name": VOD_CATEGORY["name_ar"],
            "icon": VOD_CATEGORY["icon"],
        },
    ]
    for meta in CATEGORY_META_CARDS.values():
        categories.append(
            {"id": meta["id"], "name": meta["name_ar"], "icon": meta["icon"]}
        )

    content = f"""/**
 * Marsa Store - Product Data
 * Sources:
 *  - SEAGM Card Products List
 *  - SEAGM Video on Demand Prices
 * Exchange rate: 1 USD = {RATE} MRU
 * Logos via Clearbit (logo.clearbit.com)
 * All prices in Mauritanian Ouguiya (MRU)
 * Duplicates removed by normalized product name
 *
 * أعد تشغيل: python scripts/generate_products.py
 */

const USD_TO_MRU = {RATE};

const CATEGORIES = {json.dumps(categories, ensure_ascii=False, indent=2)};

const PRODUCTS = {json.dumps(products, ensure_ascii=False, indent=2)};

const STORE_CONFIG = {{
  name: "Marsa",
  currency: "أوقية",
  currencyCode: "MRU",
  whatsappNumber: "{WHATSAPP}",
  tagline: "متجرك الموثوق للخدمات الرقمية",
  usdToMru: {RATE},
}};
"""

    OUT_PATH.write_text(content, encoding="utf-8")
    with_img = sum(1 for p in products if p.get("image"))
    print(f"Wrote {OUT_PATH}")
    print(f"Products total: {len(products)} | logos: {with_img}")
    print(f"VOD added: {vod_added}")
    print(f"Duplicates skipped: {len(skipped_dupes)}")
    cats = defaultdict(int)
    for p in products:
        cats[p["categoryName"]] += 1
    for k, v in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    if skipped_dupes[:10]:
        print("Sample skipped:", skipped_dupes[:10])


if __name__ == "__main__":
    main()
